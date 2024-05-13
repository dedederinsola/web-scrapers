from urllib.request import urlopen
from openpyxl import Workbook
from bs4 import BeautifulSoup
import re
import requests
import ssl

def scraper(url):
    context = ssl._create_unverified_context()

    wb = Workbook()

    urlpull = urlopen(url, context=context)

    target = urlpull.read().decode("utf-8")

    target = BeautifulSoup(target, "html.parser")

    links = target.find_all("a", href = re.compile(r".*/category/books/.*"))

    for link in links:
        
        ws = wb.create_sheet(title=link.get_text().strip())


        initial_url = "https://books.toscrape.com/" + link['href']

        category_url = []

        category_url.append(initial_url)

        first_iteration = True
        

        for i in range(2, 1000):
            if first_iteration:
                nextpage = re.sub("index", "page-" + str(i), category_url[-1])
                first_iteration = False
            else:
                nextpage = re.sub("page-" + str(i-1), "page-" + str(i), category_url[-1])

            response = requests.get(nextpage)  # Send a GET request to the next page URL
            if response.status_code == 200:  # Check if the response status code is 200 (OK)
                category_url.append(nextpage)  # Append the valid URL to the list
            else:
                break 


        for cat in category_url:
            category_pull = urlopen(cat, context=context)
            category_html = category_pull.read().decode("utf-8")
            category = BeautifulSoup(category_html, "html.parser")
            
                
            articles = category.find_all("article", class_="product_pod")

            for article in articles:
                # Find the <a> tag with title and extract the title text
                title_tag = article.find("a", title=True)
                title = title_tag.get("title") if title_tag else None
                
                # Find <p> tags with specific classes for price and availability
                price_tag = article.find("p", class_="price_color")
                availability_tag = article.find("p", class_="instock availability")
                
                # Extract the text content from the price and availability tags
                price = price_tag.get_text(strip=True) if price_tag else None
                availability = availability_tag.get_text(strip=True) if availability_tag else None
                
                if not ws['A1'].value:
                    ws['A1'] = 'Title'
                    ws['B1'] = 'Price'
                    ws['C1'] = 'Availability'
                
                ws.append([title, price, availability])
    return wb
            
def saveworkbook(wb):
    userpath = input("Enter the full file path (including the file name and extension): ")
    wb.save(userpath)

try:
    scraped = scraper("https://books.toscrape.com/index.html")
    saveworkbook(scraped)
except Exception as e:
    print(f"An error occurred: {e}")
